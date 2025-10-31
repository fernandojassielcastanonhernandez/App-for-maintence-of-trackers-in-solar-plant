import sys
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QWidget, QGridLayout, QPushButton,
    QVBoxLayout, QLabel, QHBoxLayout, QListWidget, QFrame,
    QComboBox, QTextEdit, QDialog, QFileDialog, QScrollArea
)
from PyQt5.QtCore import QSize, Qt
from PyQt5.QtGui import (
    QColor, QPalette, QFont, QPainter, QPainterPath
) 
import firebase_admin
from firebase_admin import credentials, db
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

# --- Configuración de Firebase con manejo de PyInstaller ---

# 1. Encontrar la ruta correcta del archivo JSON
try:
    # Ruta temporal cuando se ejecuta como ejecutable de PyInstaller
    base_path = sys._MEIPASS
except AttributeError:
    # Ruta normal cuando se ejecuta directamente desde el script
    base_path = os.path.dirname(os.path.abspath(__file__))

credentials_file_path = os.path.join(base_path, 'aldebarantrackers-firebase-adminsdk-fbsvc-1542c527bc.json')

# 2. Inicialización de Firebase
try:
    cred = credentials.Certificate(credentials_file_path)
    firebase_admin.initialize_app(cred, {
        'databaseURL': 'https://aldebarantrackers-default-rtdb.firebaseio.com/'
    })
    firebase_initialized = True
except Exception as e:
    print(f"Error al inicializar Firebase: {e}")
    firebase_initialized = False


# --- Disposición de trackers (Nombres de trackers corregidos) ---
INVERSOR_1 = [
    [     "",         "",            "",          "",    "I01-01-13", "I01-02-01", "I01-02-05", "I01-02-09", "I01-02-13", "I01-03-01", "I01-03-05", "I01-03-09", "I01-03-13", "I01-04-01", "I01-04-05", "I01-04-09", "I01-04-13", "I01-05-01", "I01-05-05", "I01-05-09", "I01-05-13", "I01-06-01", "I01-06-05", "I01-06-09", "I01-06-13"],
    ["I01-01-01", "I01-01-04", "I01-01-07", "I01-01-10", "I01-01-14", "I01-02-02", "I01-02-06", "I01-02-10", "I01-02-14", "I01-03-02", "I01-03-06", "I01-03-10", "I01-03-14", "I01-04-02", "I01-04-06", "I01-04-10", "I01-04-14", "I01-05-02", "I01-05-06", "I01-05-10", "I01-05-14", "I01-06-02", "I01-06-06", "I01-06-10", "I01-06-14"],
    ["I01-01-02", "I01-01-05", "I01-01-08", "I01-01-11", "I01-01-15", "I01-02-03", "I01-02-07", "I01-02-11", "I01-02-15", "I01-03-03", "I01-03-07", "I01-03-11", "I01-03-15", "I01-04-03", "I01-04-07", "I01-04-11", "I01-04-15", "I01-05-03", "I01-05-07", "I01-05-11", "I01-05-15", "I01-06-03", "I01-06-07", "I01-06-11", "I01-06-15"],
    ["I01-01-03", "I01-01-06", "I01-01-09", "I01-01-12", "I01-01-16", "I01-02-04", "I01-02-08", "I01-02-12", "I01-02-16", "I01-03-04", "I01-03-08", "I01-03-12", "I01-03-16", "I01-04-04", "I01-04-08", "I01-04-12", "I01-04-16", "I01-05-04", "I01-05-08", "I01-05-12", "I01-05-16", "I01-06-04", "I01-06-08", "I01-06-12", "I01-06-16"],
    ['', '', '', '', '', '', '', '', '', '', '', '', 'I01-07-01', 'I01-07-05', 'I01-07-09', 'I01-07-13', 'I01-08-01', 'I01-08-05', 'I01-08-09', 'I01-08-13', 'I01-09-01', 'I01-09-05', 'I01-09-09', 'I01-09-13', 'I01-10-01', 'I01-10-05', 'I01-10-09', 'I01-10-13', 'I01-11-01', 'I01-11-05', 'I01-11-09', 'I01-11-13', 'I01-12-01', 'I01-12-05', 'I01-12-09', 'I01-12-13'],
    ['', '', '', '', '', '', '', '', '', '', '', '', 'I01-07-02', 'I01-07-06', 'I01-07-10', 'I01-07-14', 'I01-08-02', 'I01-08-06', 'I01-08-10', 'I01-08-14', 'I01-09-02', 'I01-09-06', 'I01-09-10', 'I01-09-14', 'I01-10-02', 'I01-10-06', 'I01-10-10', 'I01-10-14', 'I01-11-02', 'I01-11-06', 'I01-11-10', 'I01-11-14', 'I01-12-02', 'I01-12-06', 'I01-12-10', 'I01-12-14'],
    ['', '', '', '', '', '', '', '', '', '', '', '', 'I01-07-03', 'I01-07-07', 'I01-07-11', 'I01-07-15', 'I01-08-03', 'I01-08-07', 'I01-08-11', 'I01-08-15', 'I01-09-03', 'I01-09-07', 'I01-09-11', 'I01-09-15', 'I01-10-03', 'I01-10-07', 'I01-10-11', 'I01-10-15', 'I01-11-03', 'I01-11-07', 'I01-11-11', 'I01-11-15', 'I01-12-03', 'I01-12-07', 'I01-12-11', 'I01-12-15'],
    ['', '', '', '', '', '', '', '', '', '', '', '', 'I01-07-04', 'I01-07-08', 'I01-07-12', 'I01-07-16', 'I01-08-04', 'I01-08-08', 'I01-08-12', 'I01-08-16', 'I01-09-04', 'I01-09-08', 'I01-09-12', 'I01-09-16', 'I01-10-04', 'I01-10-08', 'I01-10-12', 'I01-10-16', 'I01-11-04', 'I01-11-08', 'I01-11-12', 'I01-11-16', 'I01-12-04', 'I01-12-08', 'I01-12-12', 'I01-12-16'],
    ]
INVERSOR_2 = [
    # Cajas 7, 8, 9, 1, 2, 3, 4, 5 (líneas 62-69)
    ["I02-07-01", "I02-07-05", "I02-07-09", "I02-07-13", "I02-08-01", "I02-08-05", "I02-08-09", "I02-08-13", "I02-09-01", "I02-09-05", "I02-09-09", "I02-09-13", "I02-01-01", "I02-01-05", "I02-01-09", "I02-01-13", "I02-02-01", "I02-02-05", "I02-02-09", "I02-02-13", "I02-03-01", "I02-03-05", "I02-03-09", "I02-03-13", "I02-04-01", "I02-04-05", "I02-04-09", "I02-04-13", "I02-05-01", "I02-05-05", "I02-05-09", "I02-05-13"],
    ["I02-07-02", "I02-07-06", "I02-07-10", "I02-07-14", "I02-08-02", "I02-08-06", "I02-08-10", "I02-08-14", "I02-09-02", "I02-09-06", "I02-09-10", "I02-09-14", "I02-01-02", "I02-01-06", "I02-01-10", "I02-01-14", "I02-02-02", "I02-02-06", "I02-02-10", "I02-02-14", "I02-03-02", "I02-03-06", "I02-03-10", "I02-03-14", "I02-04-02", "I02-04-06", "I02-04-10", "I02-04-14", "I02-05-02", "I02-05-06", "I02-05-10", "I02-05-14"],
    ["I02-07-03", "I02-07-07", "I02-07-11", "I02-07-15", "I02-08-03", "I02-08-07", "I02-08-11", "I02-08-15", "I02-09-03", "I02-09-07", "I02-09-11", "I02-09-15", "I02-01-03", "I02-01-07", "I02-01-11", "I02-01-15", "I02-02-03", "I02-02-07", "I02-02-11", "I02-02-15", "I02-03-03", "I02-03-07", "I02-03-11", "I02-03-15", "I02-04-03", "I02-04-07", "I02-04-11", "I02-04-15", "I02-05-03", "I02-05-07", "I02-05-11", "I02-05-15"],
    ["I02-07-04", "I02-07-08", "I02-07-12", "I02-07-16", "I02-08-04", "I02-08-08", "I02-08-12", "I02-08-16", "I02-09-04", "I02-09-08", "I02-09-12", "I02-09-16", "I02-01-04", "I02-01-08", "I02-01-12", "I02-01-16", "I02-02-04", "I02-02-08", "I02-02-12", "I02-02-16", "I02-03-04", "I02-03-08", "I02-03-12", "I02-03-16", "I02-04-04", "I02-04-08", "I02-04-12", "I02-04-16", "I02-05-04", "I02-05-08", "I02-05-12", "I02-05-16"],
       # Cajas 10, 11, 12 (líneas 70-77)
    ["", "", "", "", "", "", "", "", "", "", "", "","I02-10-01", "I02-10-05", "I02-10-09", "I02-10-13", "I02-11-01", "I02-11-05", "I02-11-09", "I02-11-13", "I02-12-01", "I02-12-05", "I02-12-09", "I02-12-13"],
    ["", "", "", "", "", "", "", "", "", "", "", "","I02-10-02", "I02-10-06", "I02-10-10", "I02-10-14", "I02-11-02", "I02-11-06", "I02-11-10", "I02-11-14", "I02-12-02", "I02-12-06", "I02-12-10", "I02-12-14"],
    ["", "", "", "", "", "", "", "", "", "", "", "","I02-10-03", "I02-10-07", "I02-10-11", "I02-10-15", "I02-11-03", "I02-11-07", "I02-11-11", "I02-11-15", "I02-12-03", "I02-12-07", "I02-12-11", "I02-12-15"],
    ["", "", "", "", "", "", "", "", "", "", "", "","I02-10-04", "I02-10-08", "I02-10-12", "I02-10-16", "I02-11-04", "I02-11-08", "I02-11-12", "I02-11-16", "I02-12-04", "I02-12-08", "I02-12-12", "I02-12-16"],
    ]
INVERSOR_3 = [
    # Cajas 1 a 6
    ["", "", "", "", "", "", "", "", "", "", "", "","I03-01-01", "I03-01-05", "I03-01-09", "I03-01-13", "I03-02-01", "I03-02-05", "I03-02-09", "I03-02-13", "I03-03-01", "I03-03-05", "I03-03-09", "I03-03-13", "I03-04-01", "I03-04-05", "I03-04-09", "I03-04-13", "I03-05-01", "I03-05-05", "I03-05-09", "I03-05-13", "I03-06-01", "I03-06-05", "I03-06-09", "I03-06-13"],
    ["", "", "", "", "", "", "", "", "", "", "", "","I03-01-02", "I03-01-06", "I03-01-10", "I03-01-14", "I03-02-02", "I03-02-06", "I03-02-10", "I03-02-14", "I03-03-02", "I03-03-06", "I03-03-10", "I03-03-14", "I03-04-02", "I03-04-06", "I03-04-10", "I03-04-14", "I03-05-02", "I03-05-06", "I03-05-10", "I03-05-14", "I03-06-02", "I03-06-06", "I03-06-10", "I03-06-14"],
    ["", "", "", "", "", "", "", "", "", "", "", "","I03-01-03", "I03-01-07", "I03-01-11", "I03-01-15", "I03-02-03", "I03-02-07", "I03-02-11", "I03-02-15", "I03-03-03", "I03-03-07", "I03-03-11", "I03-03-15", "I03-04-03", "I03-04-07", "I03-04-11", "I03-04-15", "I03-05-03", "I03-05-07", "I03-05-11", "I03-05-15", "I03-06-03", "I03-06-07", "I03-06-11", "I03-06-15"],
    ["", "", "", "", "", "", "", "", "", "", "", "","I03-01-04", "I03-01-08", "I03-01-12", "I03-01-16", "I03-02-04", "I03-02-08", "I03-02-12", "I03-02-16", "I03-03-04", "I03-03-08", "I03-03-12", "I03-03-16", "I03-04-04", "I03-04-08", "I03-04-12", "I03-04-16", "I03-05-04", "I03-05-08", "I03-05-12", "I03-05-16", "I03-06-04", "I03-06-08", "I03-06-12", "I03-06-16"],
        # Cajas 7 a 12
    ["I03-07-01", "I03-07-05", "I03-07-09", "I03-07-13", "I03-08-01", "I03-08-05", "I03-08-09", "I03-08-13", "I03-09-01", "I03-09-05", "I03-09-09", "I03-09-13", "I03-10-01", "I03-10-05", "I03-10-09", "I03-10-13", "I03-11-01", "I03-11-05", "I03-11-09", "I03-11-13", "I03-12-01", "I03-12-05", "I03-12-09", "I03-12-13"],
    ["I03-07-02", "I03-07-06", "I03-07-10", "I03-07-14", "I03-08-02", "I03-08-06", "I03-08-10", "I03-08-14", "I03-09-02", "I03-09-06", "I03-09-10", "I03-09-14", "I03-10-02", "I03-10-06", "I03-10-10", "I03-10-14", "I03-11-02", "I03-11-06", "I03-11-10", "I03-11-14", "I03-12-02", "I03-12-06", "I03-12-10", "I03-12-14"],
    ["I03-07-03", "I03-07-07", "I03-07-11", "I03-07-15", "I03-08-03", "I03-08-07", "I03-08-11", "I03-08-15", "I03-09-03", "I03-09-07", "I03-09-11", "I03-09-15", "I03-10-03", "I03-10-07", "I03-10-11", "I03-10-15", "I03-11-03", "I03-11-07", "I03-11-11", "I03-11-15", "I03-12-03", "I03-12-07", "I03-12-11", "I03-12-15"],
    ["I03-07-04", "I03-07-08", "I03-07-12", "I03-07-16", "I03-08-04", "I03-08-08", "I03-08-12", "I03-08-16", "I03-09-04", "I03-09-08", "I03-09-12", "I03-09-16", "I03-10-04", "I03-10-08", "I03-10-12", "I03-10-16", "I03-11-04", "I03-11-08", "I03-11-12", "I03-11-16", "I03-12-04", "I03-12-08", "I03-12-12", "I03-12-16"],
]
INVERSOR_4 = [
    # Cajas 3 a 12
    ["I04-03-01", "I04-03-05", "I04-03-09", "I04-03-13", "I04-04-01", "I04-04-05", "I04-04-09", "I04-04-13", "I04-05-01", "I04-05-05", "I04-05-09", "I04-05-13", "I04-06-01", "I04-06-05", "I04-06-09", "I04-06-13", "I04-07-01", "I04-07-05", "I04-07-09", "I04-07-13", "I04-08-01", "I04-08-05", "I04-08-09", "I04-08-13", "I04-09-01", "I04-09-05", "I04-09-09", "I04-09-13", "I04-10-01", "I04-10-05", "I04-10-09", "I04-10-13", "I04-11-01", "I04-11-05", "I04-11-09", "I04-11-13", "I04-12-01", "I04-12-05", "I04-12-09", "I04-12-13"],
    ["I04-03-02", "I04-03-06", "I04-03-10", "I04-03-14", "I04-04-02", "I04-04-06", "I04-04-10", "I04-04-14", "I04-05-02", "I04-05-06", "I04-05-10", "I04-05-14", "I04-06-02", "I04-06-06", "I04-06-10", "I04-06-14", "I04-07-02", "I04-07-06", "I04-07-10", "I04-07-14", "I04-08-02", "I04-08-06", "I04-08-10", "I04-08-14", "I04-09-02", "I04-09-06", "I04-09-10", "I04-09-14", "I04-10-02", "I04-10-06", "I04-10-10", "I04-10-14", "I04-11-02", "I04-11-06", "I04-11-10", "I04-11-14", "I04-12-02", "I04-12-06", "I04-12-10", "I04-12-14"],
    ["I04-03-03", "I04-03-07", "I04-03-11", "I04-03-15", "I04-04-03", "I04-04-07", "I04-04-11", "I04-04-15", "I04-05-03", "I04-05-07", "I04-05-11", "I04-05-15", "I04-06-03", "I04-06-07", "I04-06-11", "I04-06-15", "I04-07-03", "I04-07-07", "I04-07-11", "I04-07-15", "I04-08-03", "I04-08-07", "I04-08-11", "I04-08-15", "I04-09-03", "I04-09-07", "I04-09-11", "I04-09-15", "I04-10-03", "I04-10-07", "I04-10-11", "I04-10-15", "I04-11-03", "I04-11-07", "I04-11-11", "I04-11-15", "I04-12-03", "I04-12-07", "I04-12-11", "I04-12-15"],
    ["I04-03-04", "I04-03-08", "I04-03-12", "I04-03-16", "I04-04-04", "I04-04-08", "I04-04-12", "I04-04-16", "I04-05-04", "I04-05-08", "I04-05-12", "I04-05-16", "I04-06-04", "I04-06-08", "I04-06-12", "I04-06-16", "I04-07-04", "I04-07-08", "I04-07-12", "I04-07-16", "I04-08-04", "I04-08-08", "I04-08-12", "I04-08-16", "I04-09-04", "I04-09-08", "I04-09-12", "I04-09-16", "I04-10-04", "I04-10-08", "I04-10-12", "I04-10-16", "I04-11-04", "I04-11-08", "I04-11-12", "I04-11-16", "I04-12-04", "I04-12-08", "I04-12-12", "I04-12-16"],
        # Cajas 1 y 2
    ["I04-02-01", "I04-02-05", "I04-02-09", "I04-02-13", "I04-01-01", "I04-01-05", "I04-01-09", "I04-01-13"],
    ["I04-02-02", "I04-02-06", "I04-02-10", "I04-02-14", "I04-01-02", "I04-01-06", "I04-01-10", "I04-01-14"],
    ["I04-02-03", "I04-02-07", "I04-02-11", "I04-02-15", "I04-01-03", "I04-01-07", "I04-01-11", "I04-01-15"],
    ["I04-02-04", "I04-02-08", "I04-02-12", "I04-02-16", "I04-01-04", "I04-01-08", "I04-01-12", "I04-01-16"],
    ]
INVERSOR_5 = [
    # Cajas 10 a 12
    ["I05-10-01", "I05-10-05", "I05-10-09", "I05-10-13", "I05-11-01", "I05-11-05", "I05-11-09", "I05-11-13", "I05-12-01", "I05-12-05", "I05-12-09", "I05-12-13"],
    ["I05-10-02", "I05-10-06", "I05-10-10", "I05-10-14", "I05-11-02", "I05-11-06", "I05-11-10", "I05-11-14", "I05-12-02", "I05-12-06", "I05-12-10", "I05-12-14"],
    ["I05-10-03", "I05-10-07", "I05-10-11", "I05-10-15", "I05-11-03", "I05-11-07", "I05-11-11", "I05-11-15", "I05-12-03", "I05-12-07", "I05-12-11", "I05-12-15"],
    ["I05-10-04", "I05-10-08", "I05-10-12", "I05-10-16", "I05-11-04", "I05-11-08", "I05-11-12", "I05-11-16", "I05-12-04", "I05-12-08", "I05-12-12", "I05-12-16"],
        # Cajas 7, 8 y 9
    ["I05-07-01", "I05-07-05", "I05-07-09", "I05-07-13", "I05-08-01", "I05-08-05", "I05-08-09", "I05-08-13", "I05-09-01", "I05-09-05", "I05-09-09", "I05-09-13"],
    ["I05-07-02", "I05-07-06", "I05-07-10", "I05-07-14", "I05-08-02", "I05-08-06", "I05-08-10", "I05-08-14", "I05-09-02", "I05-09-06", "I05-09-10", "I05-09-14"],
    ["I05-07-03", "I05-07-07", "I05-07-11", "I05-07-15", "I05-08-03", "I05-08-07", "I05-08-11", "I05-08-15", "I05-09-03", "I05-09-07", "I05-09-11", "I05-09-15"],
    ["I05-07-04", "I05-07-08", "I05-07-12", "I05-07-16", "I05-08-04", "I05-08-08", "I05-08-12", "I05-08-16", "I05-09-04", "I05-09-08", "I05-09-12", "I05-09-16"],
        # Cajas 1 a 6
    ["I05-01-01", "I05-01-05", "I05-01-09", "I05-01-13", "I05-02-01", "I05-02-05", "I05-02-09", "I05-02-13", "I05-03-01", "I05-03-05", "I05-03-09", "I05-03-13", "I05-04-01", "I05-04-05", "I05-04-09", "I05-04-13", "I05-05-01", "I05-05-05", "I05-05-09", "I05-05-13", "I05-06-01", "I05-06-05", "I05-06-09", "I05-06-13"],
    ["I05-01-02", "I05-01-06", "I05-01-10", "I05-01-14", "I05-02-02", "I05-02-06", "I05-02-10", "I05-02-14", "I05-03-02", "I05-03-06", "I05-03-10", "I05-03-14", "I05-04-02", "I05-04-06", "I05-04-10", "I05-04-14", "I05-05-02", "I05-05-06", "I05-05-10", "I05-05-14", "I05-06-02", "I05-06-06", "I05-06-10", "I05-06-14"],
    ["I05-01-03", "I05-01-07", "I05-01-11", "I05-01-15", "I05-02-03", "I05-02-07", "I05-02-11", "I05-02-15", "I05-03-03", "I05-03-07", "I05-03-11", "I05-03-15", "I05-04-03", "I05-04-07", "I05-04-11", "I05-04-15", "I05-05-03", "I05-05-07", "I05-05-11", "I05-05-15", "I05-06-03", "I05-06-07", "I05-06-11", "I05-06-15"],
    ["I05-01-04", "I05-01-08", "I05-01-12", "I05-01-16", "I05-02-04", "I05-02-08", "I05-02-12", "I05-02-16", "I05-03-04", "I05-03-08", "I05-03-12", "I05-03-16", "I05-04-04", "I05-04-08", "I05-04-12", "I05-04-16", "I05-05-04", "I05-05-08", "I05-05-12", "I05-05-16", "I05-06-04", "I05-06-08", "I05-06-12", "I05-06-16"],
    ]

INVERSORES = {
    "Inversor 1": INVERSOR_1,
    "Inversor 2": INVERSOR_2,
    "Inversor 3": INVERSOR_3,
    "Inversor 4": INVERSOR_4,
    "Inversor 5": INVERSOR_5,
}

# ----------------------------------------------------
# CLASE CORREGIDA: Dibuja fondo y texto manualmente para visibilidad
# ----------------------------------------------------
class TrackerButton(QPushButton):
    def __init__(self, name):
        self.name = name 
        super().__init__("") 
        
        self.estado = "activo"
        self.motivo = ""
        self.last_update = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.historial = []
        self.info = ""
        
        self.setFixedSize(QSize(25, 70)) 
        
        font = QFont()
        font.setPointSize(7)
        font.setBold(True)
        self.setFont(font)
        
        # Inicializa el atributo de color (no la paleta del widget)
        self.color = QColor("green") 
        self.update_color()

    def update_color(self):
        # Establece el atributo de color que paintEvent leerá
        self.color = QColor("green" if self.estado == "activo" else "red")
        self.update() # Fuerza el repintado

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        rect = self.rect()
        
        # 1. DIBUJAR SOLO EL BORDE
        painter.setPen(self.color)
        painter.setBrush(Qt.NoBrush)
        border_width = 5
        border_rect = rect.adjusted(border_width//2, border_width//2, -border_width//2, -border_width//2)
        painter.drawRoundedRect(border_rect, 5.0, 5.0)
        
       
        path = QPainterPath()
        
        # Usamos las coordenadas del rectángulo ajustado (QRect)
        adjusted_rect = rect.adjusted(1, 1, -1, -1) 
        
        # Llamamos a addRoundedRect(x, y, w, h, xRadius, yRadius)
        # Esto resuelve la ambigüedad de tipos.
        path.addRoundedRect(
            adjusted_rect.x(), 
            adjusted_rect.y(), 
            adjusted_rect.width(), 
            adjusted_rect.height(), 
            5.0, 5.0
        )
    
        painter.setPen(QColor(0, 0, 0, 0)) 
        painter.drawPath(path)
        
        # Dibuja el texto de forma vertical
        painter.save()
        painter.translate(rect.center())
        painter.rotate(90)
        font_metrics = painter.fontMetrics()
        text_width = font_metrics.width(self.name)
        text_height = font_metrics.height()
        painter.setPen(Qt.black)
        # Ajusta la posición para centrar el texto
        # Ajusta la posición horizontal sumando un desplazamiento mayor (+18)
        # Mueve el texto más hacia arriba (reduce el valor vertical)
        painter.drawText(-rect.height()//2, text_height//2 - 12, rect.height(), text_height, Qt.AlignCenter, self.name)
        painter.restore()
# ----------------------------------------------------

class DatosDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Resumen de Datos de Trackers")
        self.setFixedSize(400, 300)
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)
        
        title_label = QLabel("Estado de los Trackers")
        title_label.setStyleSheet("font-size: 20px; font-weight: bold;")
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        self.total_trackers_label = QLabel("Total de Trackers: --")
        self.total_trackers_label.setStyleSheet("font-size: 16px;")
        layout.addWidget(self.total_trackers_label)
        
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        layout.addWidget(line)

        self.activos_label = QLabel("Trackers Activos: --")
        self.activos_label.setStyleSheet("font-size: 16px; color: green; font-weight: bold;")
        layout.addWidget(self.activos_label)

        line_2 = QFrame()
        line_2.setFrameShape(QFrame.HLine)
        line_2.setFrameShadow(QFrame.Sunken)
        layout.addWidget(line_2)

        self.mantenimiento_label = QLabel("Trackers en Mantenimiento: --")
        self.mantenimiento_label.setStyleSheet("font-size: 16px; color: red; font-weight: bold;")
        layout.addWidget(self.mantenimiento_label)
        
        self.motor_label = QLabel("  - Motor: --")
        self.transmision_label = QLabel("  - Transmisión: --")
        self.tcu_label = QLabel("  - TCU: --")
        layout.addWidget(self.motor_label)
        layout.addWidget(self.transmision_label)
        layout.addWidget(self.tcu_label)

        self.download_btn = QPushButton("Descargar datos")
        self.download_btn.clicked.connect(self.descargar_datos)
        layout.addWidget(self.download_btn)
        
        self.datos_totales = {}

    def actualizar_datos(self, datos):
        self.datos_totales = datos
        self.total_trackers_label.setText(f"Total de Trackers: {datos['total']}")
        self.activos_label.setText(f"Trackers Activos: {datos['activos']}")
        self.mantenimiento_label.setText(f"Trackers en Mantenimiento: {datos['mantenimiento']}")
        self.motor_label.setText(f"  - Motor: {datos['motivos']['motor']}")
        self.transmision_label.setText(f"  - Transmisión: {datos['motivos']['transmision']}")
        self.tcu_label.setText(f"  - TCU: {datos['motivos']['TCU']}")

    def descargar_datos(self):
        if not firebase_initialized:
            print("Error: Firebase no está inicializado.")
            return

        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "Guardar Archivo Excel", "datos_trackers.xlsx", "Excel Files (*.xlsx)", options=options)
        
        if not file_path:
            return

        datos_completos = {}
        try:
            ref = db.reference('inversores')
            datos_completos = ref.get() or {}
        except Exception as e:
            print(f"Error al obtener datos completos de Firebase: {e}")
            return
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Trackers"

        headers = ["Inversor", "Tracker", "Estado", "Motivo", "Última Actualización", "Información"]
        ws.append(headers)

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for inversor_name, trackers_data in datos_completos.items():
            if trackers_data:
                for tracker_name, data in trackers_data.items():
                    estado = data.get('estado', 'activo')
                    motivo = data.get('motivo', '')
                    last_update = data.get('last_update', '')
                    # Solo dejar la fecha (YYYY-MM-DD)
                    if last_update:
                        last_update = last_update.split(' ')[0]
                    info = data.get('info', '')
                    ws.append([inversor_name, tracker_name, estado, motivo, last_update, info])

        try:
            # Auto-ajuste de columnas (simple)
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(file_path)
            print(f"Datos guardados en {file_path}")
        except Exception as e:
            print(f"Error al guardar el archivo: {e}")

class MainWindow(QWidget):
    INVERSORES = INVERSORES
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Control de Mantenimiento - Planta Solar Aldebarán")
        self.resize(1400, 750)
        layout = QHBoxLayout(self)

        left_panel_layout = QVBoxLayout()
        self.inversores = QListWidget()
        for inv in self.INVERSORES.keys():
            self.inversores.addItem(inv)
        self.inversores.currentTextChanged.connect(self.cargar_inversor)
        left_panel_layout.addWidget(self.inversores, 1)

        self.datos_btn = QPushButton("Datos")
        self.datos_btn.clicked.connect(self.mostrar_datos_totales)
        left_panel_layout.addWidget(self.datos_btn)
        
        layout.addLayout(left_panel_layout, 1)

        self.grid_frame = QFrame()
        self.grid_layout = QGridLayout(self.grid_frame)
        
        # --- CORRECCIONES DE ESPACIADO ---
        self.grid_layout.setSpacing(0) 
        self.grid_layout.setContentsMargins(0, 0, 0, 0)
        # --- FIN DE CORRECCIONES ---

    # Scroll para trackers
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scroll_area.setWidget(self.grid_frame)
        layout.addWidget(self.scroll_area, 6)

        self.detail_frame = QFrame()
        self.detail_frame.setStyleSheet("background-color: white; border: 1px solid #ccc; padding: 10px;")
        detail_layout = QVBoxLayout(self.detail_frame)

        self.detalle_label = QLabel("Selecciona un tracker")
        self.detalle_label.setAlignment(Qt.AlignCenter)
        detail_layout.addWidget(self.detalle_label)

        detail_layout.addWidget(QLabel("Estado:"))
        self.estado_combo = QComboBox()
        self.estado_combo.addItems(["activo", "mantenimiento"])
        detail_layout.addWidget(self.estado_combo)

        detail_layout.addWidget(QLabel("Motivo (si aplica):"))
        self.motivo_combo = QComboBox()
        self.motivo_combo.addItems(["", "motor", "transmision", "TCU"])
        detail_layout.addWidget(self.motivo_combo)

        self.last_update_label = QLabel("Última actualización: -")
        detail_layout.addWidget(self.last_update_label)

        detail_layout.addWidget(QLabel("Información:"))
        self.info_text = QTextEdit()
        detail_layout.addWidget(self.info_text)

        self.save_btn = QPushButton("Guardar cambios")
        self.save_btn.clicked.connect(self.guardar_cambios)
        detail_layout.addWidget(self.save_btn)

        detail_layout.addWidget(QLabel("Historial de cambios:"))
        self.historial_view = QTextEdit()
        self.historial_view.setReadOnly(True)
        detail_layout.addWidget(self.historial_view, 2)

        layout.addWidget(self.detail_frame, 1)

        self.trackers = {}
        self.selected_tracker = None
        self.current_inversor = ""
        self.inversores.setCurrentRow(0)

    def cargar_inversor(self, nombre_inv):
        self.current_inversor = nombre_inv
        # Limpiar layout
        for i in reversed(range(self.grid_layout.count())):
            widget = self.grid_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)
        self.trackers = {}
        
        tracker_data = {}
        if firebase_initialized:
            ref = db.reference(f'inversores/{self.current_inversor}')
            tracker_data = ref.get() or {}

        matriz = self.INVERSORES[nombre_inv]
        for fila, row in enumerate(matriz):
            for col, tracker_name in enumerate(row):
                if tracker_name == "":
                    continue
                
                btn = TrackerButton(tracker_name) 
                if tracker_name in tracker_data:
                    data = tracker_data[tracker_name]
                    btn.estado = data.get('estado', 'activo')
                    btn.motivo = data.get('motivo', '')
                    btn.last_update = data.get('last_update', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    btn.info = data.get('info', '')
                    btn.historial = data.get('historial', [])
                
                # Usamos update_color para establecer el color interno del botón
                btn.update_color()

                btn.clicked.connect(lambda _, b=btn: self.mostrar_detalle(b))
                self.grid_layout.addWidget(btn, fila, col)
                self.trackers[tracker_name] = btn

    def mostrar_detalle(self, tracker_btn):
        self.selected_tracker = tracker_btn
        self.detalle_label.setText(f"Tracker: {tracker_btn.name}")
        self.estado_combo.setCurrentText(tracker_btn.estado)
        self.motivo_combo.setCurrentText(tracker_btn.motivo)
        self.last_update_label.setText(f"Última actualización: {tracker_btn.last_update}")
        self.info_text.setPlainText(tracker_btn.info)
        self.historial_view.clear()
        for h in tracker_btn.historial:
            self.historial_view.append(h)

    def guardar_cambios(self):
        if not self.selected_tracker or not firebase_initialized:
            return

        tracker_btn = self.selected_tracker
        tracker_btn.estado = self.estado_combo.currentText()
        tracker_btn.motivo = self.motivo_combo.currentText()
        tracker_btn.last_update = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        tracker_btn.info = self.info_text.toPlainText()
        tracker_btn.update_color()

        log = f"{tracker_btn.last_update} | {tracker_btn.estado} | {tracker_btn.motivo}"
        tracker_btn.historial.append(log)

        data = {
            'estado': tracker_btn.estado,
            'motivo': tracker_btn.motivo,
            'last_update': tracker_btn.last_update,
            'info': tracker_btn.info,
            'historial': tracker_btn.historial
        }
        ref = db.reference(f'inversores/{self.current_inversor}/{tracker_btn.name}')
        ref.set(data)

        self.last_update_label.setText(f"Última actualización: {tracker_btn.last_update}")
        self.historial_view.clear()
        for h in tracker_btn.historial:
            self.historial_view.append(h)

    def mostrar_datos_totales(self):
        if not firebase_initialized:
            print("Firebase no está inicializado.")
            return

        datos_totales = self.obtener_datos_totales()
        dialogo = DatosDialog(self)
        dialogo.actualizar_datos(datos_totales)
        dialogo.exec_()
    
    def obtener_datos_totales(self):
        total_trackers = 960
        mantenimiento_count = 0
        motivos_count = {"motor": 0, "transmision": 0, "TCU": 0}

        try:
            ref = db.reference('inversores')
            inversores_data = ref.get()

            if inversores_data:
                for inv_key, inv_data in inversores_data.items():
                    if inv_data:
                        for tracker_name, tracker_data in inv_data.items():
                            estado = tracker_data.get('estado', 'activo')
                            motivo = tracker_data.get('motivo', '')
                            
                            if estado == 'mantenimiento':
                                mantenimiento_count += 1
                                if motivo in motivos_count:
                                    motivos_count[motivo] += 1

        except Exception as e:
            print(f"Error al obtener datos de Firebase: {e}")
            return {
                "total": total_trackers,
                "activos": 0,
                "mantenimiento": 0,
                "motivos": {"motor": 0, "transmision": 0, "TCU": 0}
            }

        activos_count = total_trackers - mantenimiento_count

        return {
            "total": total_trackers,
            "activos": activos_count,
            "mantenimiento": mantenimiento_count,
            "motivos": motivos_count
        }

if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())